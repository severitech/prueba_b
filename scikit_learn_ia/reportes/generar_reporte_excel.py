# scikit_learn_ia/reportes_excel.py
import os
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

# Rutas base ajustadas a tu estructura real
BASE_DIR = Path(__file__).resolve().parent.parent  # ← sube un nivel (sale de /reportes)
DATASET_PATH = BASE_DIR / "datasets" / "predicciones_cantidades_mensuales.csv"
OUTPUT_PATH = BASE_DIR / "datasets" / "reporte_predicciones.xlsx"



def crear_reporte_excel(desde_csv: bool = True, predicciones=None):
    """
    Crea un reporte Excel con las predicciones de demanda de productos.
    Si desde_csv=True, lee automáticamente el archivo de predicciones generado por IA.
    """
    try:
        # 🧩 Cargar predicciones desde CSV si no se pasan manualmente
        if desde_csv:
            if not DATASET_PATH.exists():
                raise FileNotFoundError(f"No se encontró el archivo: {DATASET_PATH}")
            df = pd.read_csv(DATASET_PATH)
            df = df.sort_values(["anio", "mes"])
            predicciones = df.to_dict(orient="records")

        if not predicciones:
            raise ValueError("No hay datos de predicciones disponibles.")

        # 📘 Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Predicciones de Demanda"

        # 🏷️ Encabezado principal
        ws.merge_cells("A1:F1")
        titulo = ws["A1"]
        titulo.value = "Reporte de Predicciones de Demanda de Productos"
        titulo.font = Font(bold=True, size=14, color="FFFFFF")
        titulo.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        titulo.alignment = Alignment(horizontal="center", vertical="center")

        ws.append([])  # línea vacía
        ws.append(["Año", "Mes", "Predicción", "Mínimo", "Máximo", "Confianza (%)"])

        # 🎨 Formato de encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        align_center = Alignment(horizontal="center")

        for cell in ws[3]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # 🧾 Agregar datos
        for p in predicciones:
            anio = int(p.get("anio", 0))
            mes = int(p.get("mes", 0))
            pred = int(p.get("cantidad_predicha", 0))
            min_v = int(p.get("minimo", 0))
            max_v = int(p.get("maximo", 0))
            conf = round(float(p.get("confianza", 0)) * 100, 1)

            ws.append([anio, mes, pred, min_v, max_v, conf])

        # 📊 Ajustar ancho de columnas
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 15

        # 🔲 Bordes de tabla
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for row in ws.iter_rows(min_row=3, max_col=6, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # 🕒 Pie con fecha
        ws.append([])
        ws.append([f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.cell(row=ws.max_row, column=1).font = Font(italic=True, color="888888")

        # 💾 Guardar archivo
        wb.save(OUTPUT_PATH)
        print(f"✅ Reporte Excel generado correctamente: {OUTPUT_PATH}")
        return True

    except Exception as e:
        print(f"❌ Error al generar el reporte Excel: {e}")
        return False


# Ejemplo rápido si lo ejecutas directamente:
if __name__ == "__main__":
    crear_reporte_excel()

"""
Exportadores de reportes para SmartSales365.
Soporte para PDF, Excel y JSON.
"""
import os
import json
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io


class ExportadorPDF:
    """
    Exporta reportes a formato PDF.
    """
    
    @staticmethod
    def generar_pdf_ventas(reporte_data: dict, titulo: str = "Reporte de Ventas") -> HttpResponse:
        """
        Genera un PDF para reportes de ventas.
        """
        try:
            # Crear buffer para el PDF
            buffer = io.BytesIO()
            
            # Crear documento
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Estilos
            styles = getSampleStyleSheet()
            titulo_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Centrado
            )
            
            # Elementos del documento
            elements = []
            
            # Título
            elements.append(Paragraph(titulo, titulo_style))
            elements.append(Spacer(1, 12))
            
            # Información del período
            periodo = reporte_data.get('periodo', {})
            if periodo.get('fecha_inicio') or periodo.get('fecha_fin'):
                fecha_texto = "Período: "
                if periodo.get('fecha_inicio'):
                    fecha_texto += f"Desde {periodo['fecha_inicio'].strftime('%d/%m/%Y')} "
                if periodo.get('fecha_fin'):
                    fecha_texto += f"Hasta {periodo['fecha_fin'].strftime('%d/%m/%Y')}"
                elements.append(Paragraph(fecha_texto, styles['Normal']))
                elements.append(Spacer(1, 12))
            
            # Métricas generales
            metricas = reporte_data.get('metricas_generales', {})
            if metricas:
                elements.append(Paragraph("Métricas Generales", styles['Heading2']))
                
                metricas_data = [
                    ['Total Ventas', f"${metricas.get('total_ventas', 0):,.2f}"],
                    ['Cantidad Ventas', str(metricas.get('cantidad_ventas', 0))],
                    ['Ticket Promedio', f"${metricas.get('ticket_promedio', 0):,.2f}"],
                    ['Ventas Pagadas', str(metricas.get('ventas_pagadas', 0))],
                    ['Ventas Pendientes', str(metricas.get('ventas_pendientes', 0))],
                    ['Ventas Canceladas', str(metricas.get('ventas_canceladas', 0))]
                ]
                
                tabla_metricas = Table(metricas_data, colWidths=[2*inch, 1.5*inch])
                tabla_metricas.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(tabla_metricas)
                elements.append(Spacer(1, 20))
            
            # Ventas por categoría
            ventas_categoria = reporte_data.get('ventas_por_categoria', [])
            if ventas_categoria:
                elements.append(Paragraph("Ventas por Categoría", styles['Heading2']))
                
                categoria_data = [['Categoría', 'Total Ventas', 'Cantidad Vendida', 'Productos']]
                for cat in ventas_categoria[:10]:  # Limitar a 10 categorías
                    categoria_data.append([
                        cat.get('producto__categoria__descripcion', 'N/A'),
                        f"${cat.get('total_ventas', 0):,.2f}",
                        str(cat.get('cantidad_vendida', 0)),
                        str(cat.get('cantidad_productos', 0))
                    ])
                
                tabla_categoria = Table(categoria_data, colWidths=[2*inch, 1.5*inch, 1.2*inch, 1*inch])
                tabla_categoria.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(tabla_categoria)
                elements.append(Spacer(1, 20))
            
            # Top productos
            top_productos = reporte_data.get('top_productos', [])
            if top_productos:
                elements.append(Paragraph("Top Productos", styles['Heading2']))
                
                productos_data = [['Producto', 'Categoría', 'Ventas Totales', 'Cantidad Vendida']]
                for prod in top_productos:
                    productos_data.append([
                        prod.get('producto__descripcion', 'N/A')[:30],  # Limitar longitud
                        prod.get('producto__categoria__descripcion', 'N/A'),
                        f"${prod.get('total_ventas', 0):,.2f}",
                        str(prod.get('cantidad_vendida', 0))
                    ])
                
                tabla_productos = Table(productos_data, colWidths=[2.5*inch, 1.5*inch, 1.2*inch, 1.2*inch])
                tabla_productos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(tabla_productos)
                elements.append(Spacer(1, 20))
            
            # Pie de página
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            elements.append(Paragraph("SmartSales365 - Sistema de Reportes", styles['Normal']))
            
            # Construir PDF
            doc.build(elements)
            
            # Preparar respuesta
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
            
            return response
            
        except Exception as e:
            # En caso de error, devolver respuesta de error
            return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)
    
    @staticmethod
    def generar_pdf_productos(reporte_data: dict, titulo: str = "Reporte de Productos") -> HttpResponse:
        """
        Genera un PDF para reportes de productos.
        """
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []
            
            # Título
            elements.append(Paragraph(titulo, styles['Heading1']))
            elements.append(Spacer(1, 12))
            
            # Métricas de productos
            metricas = reporte_data.get('metricas_productos', {})
            if metricas:
                elements.append(Paragraph("Resumen de Productos", styles['Heading2']))
                
                metricas_data = [
                    ['Ingresos Totales', f"${metricas.get('total_ingresos', 0):,.2f}"],
                    ['Unidades Vendidas', str(metricas.get('total_unidades_vendidas', 0))],
                    ['Productos Activos', str(metricas.get('productos_activos', 0))],
                    ['Precio Promedio', f"${metricas.get('precio_promedio', 0):,.2f}"]
                ]
                
                tabla_metricas = Table(metricas_data, colWidths=[2*inch, 1.5*inch])
                tabla_metricas.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(tabla_metricas)
                elements.append(Spacer(1, 20))
            
            # Lista de productos
            productos = reporte_data.get('productos', [])
            if productos:
                elements.append(Paragraph("Productos Más Vendidos", styles['Heading2']))
                
                productos_data = [['Producto', 'Categoría', 'Ventas', 'Cantidad', 'Clientes']]
                for prod in productos[:15]:  # Limitar a 15 productos
                    productos_data.append([
                        prod.get('producto__descripcion', 'N/A')[:25],
                        prod.get('producto__categoria__descripcion', 'N/A'),
                        f"${prod.get('ventas_totales', 0):,.2f}",
                        str(prod.get('cantidad_vendida', 0)),
                        str(prod.get('clientes_unicos', 0))
                    ])
                
                tabla_productos = Table(productos_data, colWidths=[2*inch, 1.5*inch, 1*inch, 0.8*inch, 0.8*inch])
                tabla_productos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(tabla_productos)
            
            # Pie de página
            elements.append(Spacer(1, 20))
            elements.append(Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="reporte_productos_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
            
            return response
            
        except Exception as e:
            return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)


class ExportadorExcel:
    """
    Exporta reportes a formato Excel.
    """
    
    @staticmethod
    def generar_excel_ventas(reporte_data: dict, titulo: str = "Reporte de Ventas") -> HttpResponse:
        """
        Genera un archivo Excel para reportes de ventas.
        """
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Ventas"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Título
            ws.merge_cells('A1:F1')
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = center_align
            
            # Información del período
            periodo = reporte_data.get('periodo', {})
            fecha_texto = ""
            if periodo.get('fecha_inicio'):
                fecha_texto += f"Desde: {periodo['fecha_inicio'].strftime('%d/%m/%Y')} "
            if periodo.get('fecha_fin'):
                fecha_texto += f"Hasta: {periodo['fecha_fin'].strftime('%d/%m/%Y')}"
            
            if fecha_texto:
                ws.merge_cells('A2:F2')
                ws['A2'] = fecha_texto
                ws['A2'].alignment = center_align
            
            # Métricas generales
            ws['A4'] = "Métricas Generales"
            ws['A4'].font = Font(bold=True, size=12)
            
            metricas = reporte_data.get('metricas_generales', {})
            metricas_rows = [
                ['Total Ventas', f"${metricas.get('total_ventas', 0):,.2f}"],
                ['Cantidad Ventas', metricas.get('cantidad_ventas', 0)],
                ['Ticket Promedio', f"${metricas.get('ticket_promedio', 0):,.2f}"],
                ['Ventas Pagadas', metricas.get('ventas_pagadas', 0)],
                ['Ventas Pendientes', metricas.get('ventas_pendientes', 0)],
                ['Ventas Canceladas', metricas.get('ventas_canceladas', 0)]
            ]
            
            for i, (label, value) in enumerate(metricas_rows, start=5):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
            
            # Ventas por categoría
            ws['A12'] = "Ventas por Categoría"
            ws['A12'].font = Font(bold=True, size=12)
            
            ventas_categoria = reporte_data.get('ventas_por_categoria', [])
            categoria_headers = ['Categoría', 'Total Ventas', 'Cantidad Vendida', 'Productos']
            ws.append([])  # Línea en blanco
            ws.append(categoria_headers)
            
            # Aplicar estilo a headers
            for col in range(1, len(categoria_headers) + 1):
                cell = ws.cell(row=14, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            for cat in ventas_categoria:
                ws.append([
                    cat.get('producto__categoria__descripcion', 'N/A'),
                    cat.get('total_ventas', 0),
                    cat.get('cantidad_vendida', 0),
                    cat.get('cantidad_productos', 0)
                ])
            
            # Top productos
            start_row = 15 + len(ventas_categoria) + 2
            ws[f'A{start_row}'] = "Top Productos"
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
            
            top_productos = reporte_data.get('top_productos', [])
            producto_headers = ['Producto', 'Categoría', 'Ventas Totales', 'Cantidad Vendida', 'Veces Vendido']
            ws.append([])
            ws.append(producto_headers)
            
            # Aplicar estilo a headers
            header_row = start_row + 2
            for col in range(1, len(producto_headers) + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            for prod in top_productos:
                ws.append([
                    prod.get('producto__descripcion', 'N/A'),
                    prod.get('producto__categoria__descripcion', 'N/A'),
                    prod.get('total_ventas', 0),
                    prod.get('cantidad_vendida', 0),
                    prod.get('veces_vendido', 0)
                ])
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Preparar respuesta
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            return response
            
        except Exception as e:
            return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)
    
    @staticmethod
    def generar_excel_productos(reporte_data: dict, titulo: str = "Reporte de Productos") -> HttpResponse:
        """
        Genera un archivo Excel para reportes de productos.
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Productos"
            
            # Título
            ws.merge_cells('A1:E1')
            ws['A1'] = titulo
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Métricas
            metricas = reporte_data.get('metricas_productos', {})
            ws['A3'] = "Resumen"
            ws['A3'].font = Font(bold=True)
            
            metricas_data = [
                ['Ingresos Totales', metricas.get('total_ingresos', 0)],
                ['Unidades Vendidas', metricas.get('total_unidades_vendidas', 0)],
                ['Productos Activos', metricas.get('productos_activos', 0)],
                ['Precio Promedio', metricas.get('precio_promedio', 0)]
            ]
            
            for i, (label, value) in enumerate(metricas_data, start=4):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value
            
            # Lista de productos
            productos = reporte_data.get('productos', [])
            if productos:
                ws['A8'] = "Detalle de Productos"
                ws['A8'].font = Font(bold=True)
                
                headers = ['Producto', 'Categoría', 'Precio', 'Stock', 'Ventas Totales', 'Cantidad Vendida', 'Clientes Únicos', 'Tasa Conversión']
                ws.append([])
                ws.append(headers)
                
                # Estilo headers
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=10, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                
                for prod in productos:
                    ws.append([
                        prod.get('producto__descripcion', 'N/A'),
                        prod.get('producto__categoria__descripcion', 'N/A'),
                        prod.get('producto__precio', 0),
                        prod.get('producto__stock', 0),
                        prod.get('ventas_totales', 0),
                        prod.get('cantidad_vendida', 0),
                        prod.get('clientes_unicos', 0),
                        f"{prod.get('tasa_conversion', 0):.1f}%"
                    ])
            
            # Ajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_productos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            return response
            
        except Exception as e:
            return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)


class ExportadorJSON:
    """
    Exporta reportes a formato JSON.
    """
    
    @staticmethod
    def generar_json(reporte_data: dict, nombre_archivo: str = "reporte") -> HttpResponse:
        """
        Genera un archivo JSON con los datos del reporte.
        """
        try:
            # Convertir datos a JSON
            json_data = json.dumps(reporte_data, indent=2, default=str, ensure_ascii=False)
            
            # Crear respuesta
            response = HttpResponse(json_data, content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{datetime.now().strftime("%Y%m%d_%H%M")}.json"'
            
            return response
            
        except Exception as e:
            return HttpResponse(f"Error al generar JSON: {str(e)}", status=500)


class GestorExportaciones:
    """
    Gestor principal para exportar reportes en diferentes formatos.
    """
    
    def __init__(self):
        self.exportador_pdf = ExportadorPDF()
        self.exportador_excel = ExportadorExcel()
        self.exportador_json = ExportadorJSON()
    
    def exportar_reporte(self, reporte_data: dict, formato: str, tipo_reporte: str = "ventas") -> HttpResponse:
        """
        Exporta un reporte en el formato especificado.
        
        Args:
            reporte_data: Datos del reporte
            formato: 'pdf', 'excel', 'json'
            tipo_reporte: 'ventas', 'productos', 'clientes', 'inventario'
        
        Returns:
            HttpResponse con el archivo exportado
        """
        try:
            if formato == 'pdf':
                if tipo_reporte == 'productos':
                    return self.exportador_pdf.generar_pdf_productos(reporte_data)
                else:  # ventas por defecto
                    return self.exportador_pdf.generar_pdf_ventas(reporte_data)
            
            elif formato == 'excel':
                if tipo_reporte == 'productos':
                    return self.exportador_excel.generar_excel_productos(reporte_data)
                else:  # ventas por defecto
                    return self.exportador_excel.generar_excel_ventas(reporte_data)
            
            elif formato == 'json':
                nombre_archivo = f"reporte_{tipo_reporte}"
                return self.exportador_json.generar_json(reporte_data, nombre_archivo)
            
            else:
                return HttpResponse(f"Formato no soportado: {formato}", status=400)
                
        except Exception as e:
            return HttpResponse(f"Error en exportación: {str(e)}", status=500)